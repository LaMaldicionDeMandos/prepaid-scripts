from openpyxl import load_workbook

wb = load_workbook(filename='../Base.xlsx')

for sheetname in wb.sheetnames:
    print('%s' % sheetname)
